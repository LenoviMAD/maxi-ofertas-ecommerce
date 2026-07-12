"""Genera frontend/lib/data/products.json a partir de MAESTRO.xlsx.

Uso:
    python scripts/generate-products.py [ruta-a-MAESTRO.xlsx]

Requiere Python 3.10+ y openpyxl. La salida es deterministica:
mismo Excel -> mismo JSON byte a byte (RNG sembrado con el COD).
"""

import json
import random
import sys
import unicodedata
from pathlib import Path

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "MAESTRO.xlsx"
OUT_PATH = Path(__file__).resolve().parents[1] / "frontend" / "lib" / "data" / "products.json"

TARGET = 200
STOCK_CAP = 500
FREE_SHIPPING_MIN = 40_000

# Mismo orden y mismos ids que frontend/lib/data/sucursales.ts
SUCURSAL_IDS = [
    "claypole", "villa-la-florida", "santa-rosa", "solano",
    "quilmes-oeste", "dardo-rocha", "la-plata", "bernal",
]

FAM_MAP = {
    1: "Almacén", 2: "Bebidas", 3: "Lácteos y Fiambres", 4: "Golosinas",
    5: "Perfumería", 6: "Congelados", 7: "Limpieza", 8: "Mascotas",
    9: "Hogar y Bazar", 11: "Juguetería", 12: "Electro",
}


def norm_header(value):
    """'Precio de Venta Salón C/IVA' -> 'PRECIO DE VENTA SALON C/IVA' (sin acentos)."""
    text = unicodedata.normalize("NFKD", str(value))
    return text.encode("ascii", "ignore").decode().strip().upper()


def as_number(value):
    return float(value) if isinstance(value, (int, float)) else None


def header_map(row):
    return {norm_header(v): i for i, v in enumerate(row) if v is not None}


def read_maestro(wb):
    ws = wb["MAESTRO"]
    rows = ws.iter_rows(values_only=True)
    cols = header_map(next(rows))
    out = []
    for row in rows:
        cod = row[cols["COD"]]
        desc = row[cols["DESCRIPCION"]]
        uxb = as_number(row[cols["UXB"]])
        stock = as_number(row[cols["STOCK UNI/KGS"]])
        venta = as_number(row[cols["PRECIO DE VENTA SALON C/IVA"]])
        oferta = as_number(row[cols["PRECIO OFERTA SALON C/IVA"]])  # "#N/A" -> None
        fam = row[cols["FAM"]]
        ean = row[cols["EAN 13"]]
        if cod is None or not desc or not str(desc).strip():
            continue
        if not stock or stock <= 0 or not venta or venta <= 0 or not uxb or uxb <= 0:
            continue
        if fam not in FAM_MAP:
            continue
        out.append({
            "cod": int(cod),
            "desc": str(desc).strip(),
            "uxb": int(uxb),
            "stock": int(stock),
            "venta": venta,
            "oferta": oferta if oferta is not None and 0 < oferta < venta else None,
            "fam": fam,
            "ean": str(int(ean)) if isinstance(ean, (int, float)) else None,
        })
    return out


def read_descuentos(wb):
    ws = wb["DESC X CANTIDAD"]
    rows = ws.iter_rows(values_only=True)
    cols = header_map(next(rows))
    by_cod = {}
    for row in rows:
        cod = row[cols["COD"]]
        lleva = as_number(row[cols["LLEVA"]])
        pct = as_number(row[cols["% DESC"]])
        if cod is None or lleva is None or pct is None:
            continue
        by_cod.setdefault(int(cod), []).append((int(lleva), pct))
    return by_cod


def select_products(rows):
    """Todas las ofertas validas + top stock por familia, proporcional, hasta TARGET."""
    offers = [r for r in rows if r["oferta"] is not None]
    assert len(offers) <= TARGET, f"hay {len(offers)} ofertas, mas que TARGET={TARGET}"
    pool = [r for r in rows if r["oferta"] is None]
    remaining = TARGET - len(offers)

    by_fam = {}
    for r in pool:
        by_fam.setdefault(r["fam"], []).append(r)
    total = len(pool)

    # Cuotas proporcionales al tamano de la familia (metodo del mayor resto).
    quotas, fractions, assigned = {}, [], 0
    for fam, items in sorted(by_fam.items()):
        exact = remaining * len(items) / total
        quotas[fam] = int(exact)
        assigned += int(exact)
        fractions.append((exact - int(exact), fam))
    for _, fam in sorted(fractions, reverse=True)[: remaining - assigned]:
        quotas[fam] += 1

    selected = list(offers)
    for fam, items in sorted(by_fam.items()):
        items.sort(key=lambda r: (-r["stock"], r["cod"]))
        selected.extend(items[: quotas[fam]])
    assert len(selected) == TARGET, f"esperaba {TARGET}, salieron {len(selected)}"
    return selected


def bulto_price(unit_price, uxb, descuentos):
    """Fila del mismo COD con mayor LLEVA <= UXB; sin entrada, sin descuento."""
    best = None
    for lleva, pct in descuentos or []:
        if lleva <= uxb and (best is None or lleva > best[0]):
            best = (lleva, pct)
    pct = best[1] if best else 0.0
    return round(unit_price * uxb * (1 - pct / 100))


def split_stock(cod, total):
    """Reparte `total` entre las 8 sucursales; 1-3 quedan en cero. Deterministico."""
    rng = random.Random(cod)
    zeros = set(rng.sample(SUCURSAL_IDS, rng.randint(1, 3)))
    active = [s for s in SUCURSAL_IDS if s not in zeros]
    weights = {s: rng.random() + 0.05 for s in active}
    weight_sum = sum(weights.values())
    result = {s: 0 for s in SUCURSAL_IDS}
    fractions = {}
    for s in active:
        exact = total * weights[s] / weight_sum
        result[s] = int(exact)
        fractions[s] = exact - int(exact)
    leftover = total - sum(result.values())
    for s in sorted(active, key=lambda s: -fractions[s])[:leftover]:
        result[s] += 1
    return result


def build_product(row, descuentos_by_cod):
    venta, oferta = row["venta"], row["oferta"]
    unit_price = round(oferta if oferta else venta)
    before = round(venta) if oferta else None
    discount = round((1 - oferta / venta) * 100) if oferta else 0
    # Oferta cosmetica (redondea a 0% o no baja el precio entero): se muestra
    # sin descuento, conservando el precio de oferta como precio real.
    if oferta and (discount < 1 or before <= unit_price):
        before, discount = None, 0
    bulto = bulto_price(unit_price, row["uxb"], descuentos_by_cod.get(row["cod"]))
    capped = min(row["stock"], STOCK_CAP)
    return {
        "id": str(row["cod"]),
        "name": row["desc"],
        "category": FAM_MAP[row["fam"]],
        "discountPercent": discount,
        "unitPrice": unit_price,
        "unitPriceBeforeDiscount": before,
        "bultoPrice": bulto,
        "bultoQty": row["uxb"],
        "freeShipping": bulto >= FREE_SHIPPING_MIN,
        "ean": row["ean"],
        "stockBySucursal": split_stock(row["cod"], capped),
    }


def validate(products, rows_by_cod):
    assert len(products) == TARGET, f"{len(products)} productos, esperaba {TARGET}"
    assert len({p["id"] for p in products}) == TARGET, "ids duplicados"
    for p in products:
        assert p["unitPrice"] > 0 and p["bultoPrice"] > 0 and p["bultoQty"] >= 1, p["id"]
        if p["discountPercent"] > 0:
            assert p["unitPriceBeforeDiscount"] and p["unitPriceBeforeDiscount"] > p["unitPrice"], p["id"]
        else:
            assert p["unitPriceBeforeDiscount"] is None, p["id"]
        stocks = p["stockBySucursal"]
        assert set(stocks) == set(SUCURSAL_IDS), p["id"]
        capped = min(rows_by_cod[int(p["id"])]["stock"], STOCK_CAP)
        assert sum(stocks.values()) == capped, f"{p['id']}: suma {sum(stocks.values())} != {capped}"
        assert sum(1 for v in stocks.values() if v == 0) >= 1, f"{p['id']}: sin sucursal en cero"
        assert any(v > 0 for v in stocks.values()), f"{p['id']}: todo en cero"


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    rows = read_maestro(wb)
    descuentos = read_descuentos(wb)
    selected = select_products(rows)
    products = [build_product(r, descuentos) for r in selected]
    products.sort(key=lambda p: (-p["discountPercent"], p["name"]))
    validate(products, {r["cod"]: r for r in selected})
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8", newline="\n") as f:
        json.dump(products, f, ensure_ascii=False, indent=1)
        f.write("\n")
    con_oferta = sum(1 for p in products if p["discountPercent"] > 0)
    print(f"OK: {len(products)} productos ({con_oferta} con oferta) -> {OUT_PATH}")


if __name__ == "__main__":
    main()
